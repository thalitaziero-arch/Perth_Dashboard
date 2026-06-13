"""
extract_wyscout.py
==================
Perth Azzurri — Wyscout Report Extractor
-----------------------------------------
Uso semanal: substituir o PDF novo e rodar este script.

    python extract_wyscout.py

Requisitos:
    pip install pymupdf openpyxl

Configuração (editar abaixo se necessário):
"""

import os, sys, tempfile, shutil
import fitz                                   # PyMuPDF
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment

# ──────────────────────────────────────────────────────────────
# CONFIGURAÇÃO — edite aqui se precisar mudar algo
# ──────────────────────────────────────────────────────────────

# Caminho do PDF do Wyscout (substituir pelo novo toda semana)
PDF_PATH = "Perth_SC.pdf"

# Caminho do Excel que já existe (com as abas TeamStats e PlayerStats)
XLSX_PATH = "team_stats_perth.xlsx"

# Páginas a extrair (1-based, conforme numeração do PDF)
PAGES_TO_EXTRACT = [2, 3, 4, 5, 19]

# Rótulos para cada página (aparece como cabeçalho no Excel)
PAGE_LABELS = {
    2:  "Players",
    3:  "Player Stats — General",
    4:  "Player Stats — Passing",
    5:  "Formations",
    19: "Finishing",
}

# Qualidade da imagem (1.0 = 72dpi, 2.0 = 144dpi, 2.5 = 180dpi)
RENDER_ZOOM = 2.0

# Nome da aba no Excel onde as imagens serão coladas
SHEET_NAME = "WyscoutReport"

# Cores
COLOR_NAVY  = "16284F"
COLOR_GREEN = "009246"
COLOR_WHITE = "FFFFFF"
COLOR_CREAM = "EEF0F4"

# ──────────────────────────────────────────────────────────────


def extract_pages_as_png(pdf_path, pages, zoom, out_dir):
    """Rasteriza as páginas indicadas do PDF e salva como PNG."""
    doc = fitz.open(pdf_path)
    total = len(doc)
    img_paths = {}

    for pg in pages:
        if pg < 1 or pg > total:
            print(f"  ⚠  Página {pg} fora do range (PDF tem {total} páginas) — pulando.")
            continue
        mat = fitz.Matrix(zoom, zoom)
        pix = doc[pg - 1].get_pixmap(matrix=mat)
        path = os.path.join(out_dir, f"page_{pg:02d}.png")
        pix.save(path)
        img_paths[pg] = path
        print(f"  ✓  Página {pg} extraída ({pix.width}×{pix.height}px)")

    doc.close()
    return img_paths


def insert_images_into_excel(xlsx_path, sheet_name, img_paths, pages, page_labels):
    """
    Abre o Excel existente, recria a aba WyscoutReport
    e insere as imagens com cabeçalhos coloridos.
    """
    wb = load_workbook(xlsx_path)

    # Remove aba antiga se existir
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name, 0)   # coloca como primeira aba

    # Coluna A: estreita (label de margem), Coluna B: larga (imagens)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 145

    # ── Cabeçalho do relatório ──
    ws.row_dimensions[1].height = 36
    cell = ws.cell(row=1, column=1, value="PERTH AZZURRI — WYSCOUT REPORT")
    ws.merge_cells("A1:B1")
    cell.font      = Font(name="Arial", bold=True, size=14, color=COLOR_WHITE)
    cell.fill      = PatternFill("solid", start_color=COLOR_NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 18
    hint = ws.cell(row=2, column=1,
                   value="Atualizar: substituir o PDF e rodar  python extract_wyscout.py")
    ws.merge_cells("A2:B2")
    hint.font      = Font(name="Arial", italic=True, size=9, color="666666")
    hint.fill      = PatternFill("solid", start_color=COLOR_CREAM)
    hint.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── Inserir cada página ──
    current_row = 4

    for pg in pages:
        if pg not in img_paths:
            continue

        label = page_labels.get(pg, f"Página {pg}")

        # Cabeçalho da página
        ws.row_dimensions[current_row].height = 22
        lbl = ws.cell(row=current_row, column=1, value=f"  Pág {pg} · {label}")
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row,   end_column=2
        )
        lbl.font      = Font(name="Arial", bold=True, size=11, color=COLOR_WHITE)
        lbl.fill      = PatternFill("solid", start_color=COLOR_GREEN)
        lbl.alignment = Alignment(horizontal="left", vertical="center")
        current_row += 1

        # Imagem
        img_obj = XLImage(img_paths[pg])
        # Manter proporção; largura ~1050px para caber na col B
        orig_w, orig_h = img_obj.width, img_obj.height
        target_w = 1050
        scale    = target_w / orig_w
        img_obj.width  = int(orig_w * scale)
        img_obj.height = int(orig_h * scale)

        img_obj.anchor = f"B{current_row}"
        ws.add_image(img_obj)

        # Reservar linhas para a imagem (altura em pts; 1px ≈ 0.75pt; row default 15pt)
        row_count = int((img_obj.height * 0.75) / 15) + 1
        for r in range(current_row, current_row + row_count):
            ws.row_dimensions[r].height = 15

        current_row += row_count + 2     # espaço entre páginas

    wb.save(xlsx_path)
    print(f"\n  ✅  Excel salvo: {xlsx_path}")
    print(f"  📋  Aba '{sheet_name}' atualizada com {len(img_paths)} páginas.")


def main():
    print("=" * 55)
    print("  Perth Azzurri — Wyscout PDF → Excel")
    print("=" * 55)

    # Verifica arquivos
    if not os.path.exists(PDF_PATH):
        print(f"\n❌  PDF não encontrado: {PDF_PATH}")
        print("    Coloque o PDF do Wyscout na mesma pasta e tente novamente.")
        sys.exit(1)

    if not os.path.exists(XLSX_PATH):
        print(f"\n❌  Excel não encontrado: {XLSX_PATH}")
        sys.exit(1)

    print(f"\n  📄  PDF: {PDF_PATH}")
    print(f"  📊  Excel: {XLSX_PATH}")
    print(f"  📑  Páginas: {PAGES_TO_EXTRACT}\n")

    # Pasta temporária para as imagens
    tmp_dir = tempfile.mkdtemp(prefix="perth_wyscout_")

    try:
        print("Extraindo páginas do PDF...")
        img_paths = extract_pages_as_png(PDF_PATH, PAGES_TO_EXTRACT, RENDER_ZOOM, tmp_dir)

        print("\nInserindo imagens no Excel...")
        insert_images_into_excel(
            XLSX_PATH, SHEET_NAME,
            img_paths, PAGES_TO_EXTRACT, PAGE_LABELS
        )

    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)

    print("\n  Pronto! Abra o Excel e veja a aba 'WyscoutReport'.")
    print("=" * 55)


if __name__ == "__main__":
    main()
