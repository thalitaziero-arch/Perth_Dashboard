#!/usr/bin/env python3
"""
atualizar_painel.py
====================

Script de atualização semanal do painel Perth Azzurri.

COMO USAR (toda semana):
  1. Substitua o arquivo `team_stats_perth.xlsx` pelo novo Excel
     (mesma aba "TeamStats" e "DashboardData", mesmo nome de arquivo).
  2. Substitua o arquivo `NPL_Comparison.pdf` pelo novo PDF "Season Report"
     baixado do Hudl Wyscout (mesmo nome de arquivo).
  3. Substitua o arquivo `Perth_SC.pdf` pelo novo relatório do time
     (Wyscout "Team Report" do Perth, com o filtro de período em
     "todas as partidas / temporada completa" — NÃO "últimas 5 partidas",
     que é o padrão do Wyscout). Mesmo nome de arquivo.
     É dele que saem: as duas figuras de "Shots" (mapa no gol e no campo)
     e as estatísticas de cada jogador (gols, passes, duelos, etc.).
     Se o PDF estiver no filtro errado, o script detecta e avisa, sem
     aplicar os dados.
  4. Rode:
         python3 atualizar_painel.py
     Isso já publica automaticamente no GitHub (o site online atualiza
     em 1-2 minutos). Não precisa rodar git add/commit/push manualmente.
  5. Abra `perth_azzurri_painel.html` no navegador e confira.

O script:
  - Lê a aba "TeamStats" do Excel, detecta o(s) jogo(s) novo(s) que ainda
    não estão na aba "DashboardData" e os adiciona automaticamente
    (faz backup do Excel antes: team_stats_perth.bak.xlsx).
  - Roda o extrator no NPL_Comparison.pdf novo e atualiza a classificação
    e estatísticas comparativas (npl_comparison_data.json).
  - Injeta tudo (DATA, SD_TOTAL, REPORT_DATA) dentro do
    perth_azzurri_painel.html (faz backup antes: perth_azzurri_painel.bak.html).
  - Valida o JSON gerado antes de salvar — se algo der erro, o HTML
    original não é tocado.

Requisitos: pip install pandas openpyxl pdfplumber --break-system-packages
"""

import json
import math
import re
import shutil
import subprocess
import sys
from pathlib import Path

import base64
import datetime

import pandas as pd

import extract_wyscout as ew

BASE_DIR = Path(__file__).parent
EXCEL_FILE = BASE_DIR / "team_stats_perth.xlsx"
HTML_FILE = BASE_DIR / "perth_azzurri_painel.html"
NPL_PDF = BASE_DIR / "NPL_Comparison.pdf"
NPL_JSON = BASE_DIR / "npl_comparison_data.json"
EXTRACT_SCRIPT = BASE_DIR / "extract_wyscout.py"
PERTH_PDF = BASE_DIR / "Perth_SC.pdf"
GPS_FILE = BASE_DIR / "Perth_SC_GPS_Weekly_Update_v3.xlsx"
GPS_HISTORY_FILE = BASE_DIR / "gps_history.json"

TS_COLS = {
    "Date": 0, "Match": 1, "Competition": 2, "Duration": 3, "Team": 4, "Scheme": 5,
    "Goals": 6, "xG": 7, "Shots": 8, "SOT": 9, "Shots_acc_pct": 10,
    "Possession_pct": 11,
    "Losses_total": 12, "Losses_low": 13, "Losses_medium": 14, "Losses_high": 15,
    "Rec_total": 16, "Rec_low": 17, "Rec_mid": 18, "Rec_high": 19,
    "Duels_total": 20, "Duels_won": 21, "Duels_won_pct": 22,
    "Fwd_passes_pct": 63,
}


def step1_update_excel():
    print("1) Verificando jogos novos no Excel...")
    ts = pd.read_excel(EXCEL_FILE, sheet_name="TeamStats", header=None)
    dd = pd.read_excel(EXCEL_FILE, sheet_name="DashboardData")

    data_rows = ts.iloc[3:].copy()  # pula as 3 linhas de cabeçalho/médias
    existing_matches = set(dd["Match"].unique())

    new_matches = []
    for match in data_rows[TS_COLS["Match"]].unique():
        if pd.isna(match):
            continue
        if match not in existing_matches:
            new_matches.append(match)

    if not new_matches:
        print("   Nenhum jogo novo encontrado. DashboardData já está atualizado.")
        return dd

    def build_row(row, round_num):
        g = lambda c: row.iloc[c]
        return {
            "Round": round_num,
            "Date": str(g(TS_COLS["Date"]))[:10],
            "Match": g(TS_COLS["Match"]),
            "Competition": g(TS_COLS["Competition"]),
            "Duration": int(g(TS_COLS["Duration"])),
            "Team": g(TS_COLS["Team"]),
            "Scheme": g(TS_COLS["Scheme"]),
            "Goals": int(g(TS_COLS["Goals"])),
            "xG": round(float(g(TS_COLS["xG"])), 2),
            "Shots": int(g(TS_COLS["Shots"])),
            "Shots_on_target": int(g(TS_COLS["SOT"])),
            "Shots_acc_pct": round(float(g(TS_COLS["Shots_acc_pct"])), 2),
            "Passes": 0,
            "Passes_accurate": 0,
            "Passes_acc_pct": round(float(g(TS_COLS["Fwd_passes_pct"])), 2),
            "Possession_pct": round(float(g(TS_COLS["Possession_pct"])), 2),
            "Losses_total": int(g(TS_COLS["Losses_total"])),
            "Losses_low": int(g(TS_COLS["Losses_low"])),
            "Losses_medium": int(g(TS_COLS["Losses_medium"])),
            "Losses_high": int(g(TS_COLS["Losses_high"])),
            "Recoveries_total": int(g(TS_COLS["Rec_total"])),
            "Recoveries_low": int(g(TS_COLS["Rec_low"])),
            "Recoveries_medium": int(g(TS_COLS["Rec_mid"])),
            "Recoveries_high": int(g(TS_COLS["Rec_high"])),
            "Duels_total": int(g(TS_COLS["Duels_total"])),
            "Duels_won": int(g(TS_COLS["Duels_won"])),
            "Duels_won_pct": round(float(g(TS_COLS["Duels_won_pct"])), 2),
        }

    next_round = int(dd["Round"].max()) + 1
    new_rows = []
    for match in new_matches:
        rows_for_match = data_rows[data_rows[TS_COLS["Match"]] == match]
        for _, row in rows_for_match.iterrows():
            new_rows.append(build_row(row, next_round))
        next_round += 1

    print(f"   Jogo(s) novo(s) encontrado(s): {new_matches}")
    shutil.copy(EXCEL_FILE, EXCEL_FILE.with_suffix(".bak.xlsx"))
    combined = pd.concat([dd, pd.DataFrame(new_rows)], ignore_index=True)
    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        combined.to_excel(writer, sheet_name="DashboardData", index=False)
    print(f"   Excel atualizado (backup em {EXCEL_FILE.with_suffix('.bak.xlsx').name}).")
    return combined


MESES_PT = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
            "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]


def step2_update_npl_json(dashboard_df):
    print("2) Atualizando dados da NPL Comparison a partir do PDF...")
    if not NPL_PDF.exists():
        print(f"   Aviso: {NPL_PDF.name} não encontrado, pulando esta etapa.")
        return

    last_round = int(dashboard_df["Round"].max())
    today = datetime.date.today()
    round_label = f"Round {last_round} · {MESES_PT[today.month - 1]} {today.year}"

    result = subprocess.run(
        [sys.executable, str(EXTRACT_SCRIPT), str(NPL_PDF), "--round", round_label],
        capture_output=True, text=True,
    )
    print(result.stdout)
    if result.returncode != 0:
        print(result.stderr)
        print("   ATENÇÃO: erro ao processar o PDF. Confira manualmente o npl_comparison_data.json.")
    print("   IMPORTANTE: confira no resumo acima se algum aviso aparece (parser pode")
    print("   colar números errados em colunas estreitas, ex: pontos da classificação).")


def _ratio(s):
    """'11 / 5 45%' -> (11, 5, 45.0); '45 / 13' -> (45, 13, None); '-' -> (0, 0, None)."""
    if not s or s == "-":
        return 0, 0, None
    m = re.match(r"(-?\d+)\s*/\s*(-?\d+)\s*(?:(\d+(?:\.\d+)?)%)?", s)
    if not m:
        return 0, 0, None
    a, b, pct = m.groups()
    return int(a), int(b), (float(pct) if pct else None)


def _single(s):
    if not s or s == "-":
        return 0
    m = re.match(r"-?\d+(\.\d+)?", s)
    return float(m.group()) if m and "." in m.group() else int(m.group()) if m else 0


SNAPSHOT_FILE = BASE_DIR / "perth_sc_last5_snapshot.json"
APPLIED_LOG = BASE_DIR / "perth_sc_applied_matches.json"

# Fields tracked as (numerator, denominator) pairs in Table A/B/C, so they can
# be added/subtracted numerically when isolating a single match's contribution
# out of a "last N matches" rolling report.
RATIO_FIELDS_A = ["shots", "passes", "crosses", "dribbles", "duels"]
PLAIN_PAIR_FIELDS_A = ["losses", "recoveries", "cards"]  # "a/b", no pct
RATIO_FIELDS_B = ["def_duels", "off_duels", "aerial", "loose"]
RATIO_FIELDS_C = ["forward", "back", "lateral", "short_med", "long",
                  "progressive", "final_third", "through"]


def _numeric_record(rec):
    """Convert one player's raw Table A/B/C strings into plain numeric deltas
    (ints/floats only), so two windows ('last 5' snapshots) can be subtracted."""
    out = {}
    a = rec.get("A")
    if a:
        out["total_min"] = _single(a["total_min"])
        gp = a["goals_xg"].split("/") if a["goals_xg"] != "-" else ["0", "0"]
        ap = a["assists_xa"].split("/") if a["assists_xa"] != "-" else ["0", "0"]
        out["goals"] = int(gp[0].strip())
        out["xg"] = float(gp[1].strip()) if len(gp) > 1 else 0.0
        out["assists"] = int(ap[0].strip())
        out["xa"] = float(ap[1].strip()) if len(ap) > 1 else 0.0
        for f in RATIO_FIELDS_A:
            n, d, _ = _ratio(a[f])
            out[f + "_n"], out[f + "_d"] = n, d
        for f in PLAIN_PAIR_FIELDS_A:
            n, d, _ = _ratio(a[f])
            out[f + "_n"], out[f + "_d"] = n, d
        out["touches_pa"] = int(_single(a["touches_pa"]))
    b = rec.get("B")
    if b:
        for f in RATIO_FIELDS_B:
            n, d, _ = _ratio(b[f])
            out[f + "_n"], out[f + "_d"] = n, d
    c = rec.get("C")
    if c:
        for f in RATIO_FIELDS_C:
            n, d, _ = _ratio(c[f])
            out[f + "_n"], out[f + "_d"] = n, d
        out["deep_completions"] = int(_single(c["deep_completions"]))
        out["key_passes"] = int(_single(c["key_passes"]))
        out["shot_assists"] = int(_single(c["shot_assists"]))
    return out


def _subtract_numeric(new, old):
    return {k: new.get(k, 0) - old.get(k, 0) for k in new}


def _apply_delta_to_player(player, delta):
    """Add one match's numeric contribution onto a player's existing season
    totals (which are stored as compact 'n/d' strings + separate _pct field)."""
    player["matches"] = player.get("matches", 0) + 1
    player["total_min"] = player.get("total_min", 0) + delta.get("total_min", 0)
    if player["matches"]:
        player["avg_min"] = round(player["total_min"] / player["matches"])
    player["goals"] = player.get("goals", 0) + delta.get("goals", 0)
    player["xg"] = round(player.get("xg", 0.0) + delta.get("xg", 0.0), 2)
    player["assists"] = player.get("assists", 0) + delta.get("assists", 0)
    player["xa"] = round((player.get("xa") or 0.0) + delta.get("xa", 0.0), 2)

    def merge_ratio(field_str_key, field_pct_key, dn, dd):
        n0, d0, _ = _ratio((player.get(field_str_key) or "0/0").replace("/", " / "))
        n, d = n0 + dn, d0 + dd
        player[field_str_key] = f"{n}/{d}"
        player[field_pct_key] = round(d / n * 100, 1) if n else None

    for f in RATIO_FIELDS_A:
        merge_ratio(f, f + "_pct", delta.get(f + "_n", 0), delta.get(f + "_d", 0))

    player["losses_total"] = player.get("losses_total", 0) + delta.get("losses_n", 0)
    player["losses_own"] = player.get("losses_own", 0) + delta.get("losses_d", 0)
    player["recoveries"] = player.get("recoveries", 0) + delta.get("recoveries_n", 0)
    player["recoveries_opp"] = player.get("recoveries_opp", 0) + delta.get("recoveries_d", 0)
    player["touches_pa"] = player.get("touches_pa", 0) + delta.get("touches_pa", 0)
    player["yc"] = player.get("yc", 0) + delta.get("cards_n", 0)
    player["rc"] = player.get("rc", 0) + delta.get("cards_d", 0)

    dd2 = player.setdefault("duels_detail", {})
    for f in RATIO_FIELDS_B:
        n0, d0, _ = _ratio((dd2.get(f) or "0 / 0"))
        n, d = n0 + delta.get(f + "_n", 0), d0 + delta.get(f + "_d", 0)
        dd2[f] = f"{n} / {d}"
        pct_key = "def_pct" if f == "def_duels" else "off_pct" if f == "off_duels" else f + "_pct"
        dd2[pct_key] = f"{round(d / n * 100)}%" if n else "-"

    pp = player.setdefault("passing", {})
    for f in RATIO_FIELDS_C:
        n0, d0, _ = _ratio((pp.get(f) or "0 / 0"))
        n, d = n0 + delta.get(f + "_n", 0), d0 + delta.get(f + "_d", 0)
        pp[f] = f"{n} / {d}"
        pp[f + "_pct"] = round(d / n * 100, 1) if n else 0.0
    pp["deep_completions"] = pp.get("deep_completions", 0) + delta.get("deep_completions", 0)
    pp["key_passes"] = pp.get("key_passes", 0) + delta.get("key_passes", 0)
    pp["shot_assists"] = pp.get("shot_assists", 0) + delta.get("shot_assists", 0)
    return player


def step2c_update_players(dashboard_df):
    print("2c) Atualizando estatísticas de jogadores (Perth_SC.pdf)...")
    html = HTML_FILE.read_text(encoding="utf-8")
    m = re.search(r"const PLAYERS = (\[.*?\]);", html, re.S)
    base_players = {p["name"]: p for p in json.loads(m.group(1))}

    if not PERTH_PDF.exists():
        print(f"   Aviso: {PERTH_PDF.name} não encontrado, mantendo estatísticas de jogadores como estão.")
        return list(base_players.values())

    names = set(base_players)
    matches_by_name = ew.parse_perth_sc_overview(PERTH_PDF, names)
    current_round = int(dashboard_df["Round"].max())
    max_matches_in_pdf = max(matches_by_name.values()) if matches_by_name else 0
    pdf_data = ew.parse_perth_sc_pdf(PERTH_PDF, names)

    full_season = max_matches_in_pdf >= current_round - 2
    single_match = max_matches_in_pdf == 1 and not full_season
    snapshot = json.loads(SNAPSHOT_FILE.read_text()) if SNAPSHOT_FILE.exists() else None

    players = []
    if single_match:
        # Perth_SC.pdf filtered to exactly one match (e.g. by date range in
        # Wyscout) — the parsed numbers are already that match's contribution,
        # no diffing needed; just add them onto the existing season totals.
        # Guard against double-counting: identify the match (from the cover
        # page) and skip if it was already applied in a previous run.
        import fitz
        doc = fitz.open(str(PERTH_PDF))
        match_signature = doc[0].get_text().strip().splitlines()
        match_signature = match_signature[2] if len(match_signature) > 2 else doc[0].get_text().strip()
        doc.close()
        applied = json.loads(APPLIED_LOG.read_text()) if APPLIED_LOG.exists() else []

        if match_signature in applied:
            print(f"   Esse jogo ('{match_signature}') já foi somado aos totais antes "
                  f"(visto em execução anterior) — não vou somar de novo. Estatísticas "
                  f"de jogadores mantidas como estão.")
            players = list(base_players.values())
        else:
            updated = 0
            for name, base in base_players.items():
                player = dict(base)
                if name in pdf_data and not base.get("is_goalkeeper"):
                    delta = _numeric_record(pdf_data[name])
                    player = _apply_delta_to_player(player, delta)
                    updated += 1
                elif base.get("is_goalkeeper") and matches_by_name.get(name):
                    player["matches"] = player.get("matches", 0) + 1
                players.append(player)
            applied.append(match_signature)
            APPLIED_LOG.write_text(json.dumps(applied, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"   {updated} jogadores atualizados com os dados deste jogo único "
                  f"(rodada {current_round}, '{match_signature}').")
    elif full_season:
        # the PDF already covers (close to) the whole season — use it as the
        # direct source of truth, same as before.
        for name, base in base_players.items():
            player = dict(base)
            rec = pdf_data.get(name)
            matches = matches_by_name.get(name)
            if matches is not None:
                player["matches"] = matches
            if rec and not base.get("is_goalkeeper"):
                num = _numeric_record(rec)
                player["total_min"] = num.get("total_min", player.get("total_min", 0))
                if player.get("matches"):
                    player["avg_min"] = round(player["total_min"] / player["matches"])
                player["goals"] = num.get("goals", 0)
                player["xg"] = round(num.get("xg", 0.0), 2)
                player["assists"] = num.get("assists", 0)
                player["xa"] = round(num.get("xa", 0.0), 2)
                for f in RATIO_FIELDS_A:
                    n, d = num.get(f + "_n", 0), num.get(f + "_d", 0)
                    player[f] = f"{n}/{d}"
                    player[f + "_pct"] = round(d / n * 100, 1) if n else None
                player["losses_total"], player["losses_own"] = num.get("losses_n", 0), num.get("losses_d", 0)
                player["recoveries"], player["recoveries_opp"] = num.get("recoveries_n", 0), num.get("recoveries_d", 0)
                player["touches_pa"] = num.get("touches_pa", 0)
                player["yc"], player["rc"] = num.get("cards_n", 0), num.get("cards_d", 0)
                dd2 = player.setdefault("duels_detail", {})
                for f in RATIO_FIELDS_B:
                    n, d = num.get(f + "_n", 0), num.get(f + "_d", 0)
                    dd2[f] = f"{n} / {d}"
                    pct_key = "def_pct" if f == "def_duels" else "off_pct" if f == "off_duels" else f + "_pct"
                    dd2[pct_key] = f"{round(d / n * 100)}%" if n else "-"
                pp = player.setdefault("passing", {})
                for f in RATIO_FIELDS_C:
                    n, d = num.get(f + "_n", 0), num.get(f + "_d", 0)
                    pp[f] = f"{n} / {d}"
                    pp[f + "_pct"] = round(d / n * 100, 1) if n else 0.0
                pp["deep_completions"] = num.get("deep_completions", 0)
                pp["key_passes"] = num.get("key_passes", 0)
                pp["shot_assists"] = num.get("shot_assists", 0)
            players.append(player)
        print(f"   {len(players)} jogadores atualizados a partir do Perth_SC.pdf "
              f"(temporada completa, rodada {current_round}).")
    elif snapshot:
        # rolling-window report ('last N matches'). Diff against last week's
        # saved window to isolate just the newest match, then add that to the
        # existing season totals.
        updated, skipped = 0, []
        for name, base in base_players.items():
            player = dict(base)
            matches_now = matches_by_name.get(name)
            matches_before = snapshot.get("matches", {}).get(name)
            if (matches_now is not None and matches_before is not None
                    and matches_now == matches_before + 1 and not base.get("is_goalkeeper")):
                new_num = _numeric_record(pdf_data.get(name, {}))
                old_num = _numeric_record(snapshot.get("records", {}).get(name, {}))
                delta = _subtract_numeric(new_num, old_num)
                player = _apply_delta_to_player(player, delta)
                updated += 1
            elif matches_now is not None and matches_before is not None and matches_now != matches_before:
                skipped.append(name)
            players.append(player)
        print(f"   {updated} jogadores atualizados com o jogo novo (diferença vs. o retrato salvo).")
        if skipped:
            print(f"   AVISO: variação de partidas inesperada para {skipped} — não atualizados, "
                  f"confira manualmente.")
    else:
        print(f"   AVISO: Perth_SC.pdf cobre só {max_matches_in_pdf} jogo(s) (rodada atual: "
              f"{current_round}) e ainda não há um retrato salvo de uma semana anterior para "
              f"comparar — não é possível isolar só o jogo novo ainda. Estatísticas de "
              f"jogadores mantidas como estão. Um retrato deste PDF foi salvo agora; na "
              f"próxima atualização semanal (com o próximo PDF de últimas partidas) já será "
              f"possível extrair só o jogo novo automaticamente.")
        players = list(base_players.values())

    if not single_match:
        # only "last N matches" / full-season windows are useful as a future
        # diff baseline; a single-match report has a different shape.
        SNAPSHOT_FILE.write_text(json.dumps(
            {"matches": matches_by_name, "records": pdf_data}, ensure_ascii=False, indent=2
        ), encoding="utf-8")
    return players


def step2b_update_shots_images():
    print("2b) Recortando as figuras 'Shots' do relatório do último jogo (Perth_SC.pdf)...")
    if not PERTH_PDF.exists():
        print(f"   Aviso: {PERTH_PDF.name} não encontrado, pulando esta etapa.")
        return None, None
    import fitz  # pymupdf
    doc = fitz.open(str(PERTH_PDF))
    pg = None
    for i in range(len(doc)):
        text = doc[i].get_text()
        if text.startswith("Shots\nGoal\nOn target\nMiss"):
            pg = i + 1
            break
    doc.close()
    if pg is None:
        print("   Aviso: não encontrei a página 'Shots' nesse PDF, pulando as figuras.")
        return None, None

    def crop(page_num, x0, y0, x1, y1, scale=4.5):
        doc = fitz.open(str(PERTH_PDF))
        page = doc[page_num - 1]
        mat = fitz.Matrix(scale, scale)
        pix = page.get_pixmap(matrix=mat, clip=fitz.Rect(x0, y0, x1, y1))
        png = pix.tobytes("png")
        doc.close()
        return base64.b64encode(png).decode()

    img1 = crop(pg, x0=0, y0=55, x1=298, y1=245)
    img2 = crop(pg, x0=0, y0=245, x1=298, y1=475)
    print(f"   Figuras extraídas da página {pg}.")
    return img1, img2


def step2d_update_npl_images():
    print("2d) Recortando figuras 'Goals', 'Attacks by flanks' e 'Corners distribution' (NPL_Comparison.pdf)...")
    if not NPL_PDF.exists():
        print(f"   Aviso: {NPL_PDF.name} não encontrado, pulando esta etapa.")
        return None, None, None
    import fitz  # pymupdf

    def find_page(marker):
        doc = fitz.open(str(NPL_PDF))
        pg = None
        for i in range(len(doc)):
            if marker in doc[i].get_text():
                pg = i + 1
                break
        doc.close()
        return pg

    def crop(page_num, x0, y0, x1, y1, scale=4.5):
        doc = fitz.open(str(NPL_PDF))
        page = doc[page_num - 1]
        mat = fitz.Matrix(scale, scale)
        pix = page.get_pixmap(matrix=mat, clip=fitz.Rect(x0, y0, x1, y1))
        png = pix.tobytes("png")
        doc.close()
        return base64.b64encode(png).decode()

    pg_goals = find_page("Attacks by flanks")  # same page as 'Goals'
    img_flanks = img_goals = None
    if pg_goals:
        img_flanks = crop(pg_goals, x0=0, y0=320, x1=298, y1=450)
        img_goals = crop(pg_goals, x0=297, y0=320, x1=595, y1=450)
        print(f"   Figuras 'Attacks by flanks' e 'Goals' extraídas da página {pg_goals}.")
    else:
        print("   Aviso: não encontrei a página 'Attacks by flanks' / 'Goals'.")

    pg_corners = find_page("Corners distribution")
    img_corners = None
    if pg_corners:
        img_corners = crop(pg_corners, x0=5, y0=65, x1=300, y1=215)
        print(f"   Figura 'Corners distribution' extraída da página {pg_corners}.")
    else:
        print("   Aviso: não encontrei a página 'Corners distribution'.")

    return img_flanks, img_goals, img_corners


def step3_inject_html(dashboard_df, img1=None, img2=None, players=None, gps=None,
                       npl_flanks_img=None, npl_goals_img=None, npl_corners_img=None):
    print("3) Gravando dados no perth_azzurri_painel.html...")
    shutil.copy(HTML_FILE, HTML_FILE.with_suffix(".bak.html"))

    records = dashboard_df.to_dict("records")

    def clean(v):
        if isinstance(v, float) and math.isnan(v):
            return None
        return v

    records = [{k: clean(v) for k, v in r.items()} for r in records]
    for r in records:
        r["Date"] = str(r["Date"])[:10]

    html = HTML_FILE.read_text(encoding="utf-8")
    html = re.sub(r"const DATA = \[.*?\];", "const DATA = " + json.dumps(records) + ";", html, flags=re.S)

    perth_rows = [r for r in records if r["Team"] == "Perth"]
    last_round_row = max(perth_rows, key=lambda r: r["Round"])
    last_round_num = last_round_row["Round"]
    last_date = datetime.date.fromisoformat(last_round_row["Date"])
    date_label = f"{last_date.day} {MESES_PT[last_date.month - 1]} {last_date.year}"
    html = re.sub(
        r'(id="homeRoundBadge"[^>]*>)Round \d+ · WA NPLW \d+(<)',
        rf"\g<1>Round {last_round_num} · WA NPLW {last_date.year}\g<2>",
        html,
    )
    html = re.sub(
        r'(id="homeUpdatedBadge"[^>]*>)Last updated: [^<]+(<)',
        rf"\g<1>Last updated: {date_label}\g<2>",
        html,
    )

    sd_shots = sum(r["Shots"] for r in perth_rows)
    sd_sot = sum(r["Shots_on_target"] for r in perth_rows)
    sd_pct = round(sd_sot / sd_shots * 100, 1) if sd_shots else 0
    sd_xg = round(sum(r["xG"] for r in perth_rows), 2)
    sd_goals = sum(r["Goals"] for r in perth_rows)
    sd_total = f'{{"shots":{sd_shots},"on_target":{sd_sot},"pct":{sd_pct},"xg":{sd_xg},"goals":{sd_goals}}}'
    html = re.sub(r"const SD_TOTAL = \{.*?\};", f"const SD_TOTAL = {sd_total};", html, flags=re.S)

    if NPL_JSON.exists():
        npl = json.loads(NPL_JSON.read_text(encoding="utf-8"))
        html = re.sub(
            r"const REPORT_DATA = \{.*?\};\n",
            "const REPORT_DATA = " + json.dumps(npl, ensure_ascii=False) + ";\n",
            html, flags=re.S,
        )
        rankings = npl.get("playerRankings")
        if rankings:
            html = re.sub(
                r"const NPL_RANKINGS = \{.*?\n(?=\s*const playerRankings = NPL_RANKINGS\[player\.name\];)",
                "const NPL_RANKINGS = " + json.dumps(rankings, ensure_ascii=False) + ";\n  ",
                html, flags=re.S,
            )

    if img1:
        html = re.sub(r'const FINISHING_IMG1 = "[^"]*";', f'const FINISHING_IMG1 = "{img1}";', html)
    if img2:
        html = re.sub(r'const FINISHING_IMG2 = "[^"]*";', f'const FINISHING_IMG2 = "{img2}";', html)
    if npl_flanks_img:
        html = re.sub(r'const NPL_FLANKS_IMG = "[^"]*";', f'const NPL_FLANKS_IMG = "{npl_flanks_img}";', html)
    if npl_goals_img:
        html = re.sub(r'const NPL_GOALS_IMG = "[^"]*";', f'const NPL_GOALS_IMG = "{npl_goals_img}";', html)
    if npl_corners_img:
        html = re.sub(r'const NPL_CORNERS_IMG = "[^"]*";', f'const NPL_CORNERS_IMG = "{npl_corners_img}";', html)

    if players is not None:
        html = re.sub(
            r"const PLAYERS = \[.*?\];",
            "const PLAYERS = " + json.dumps(players, ensure_ascii=False) + ";",
            html, flags=re.S,
        )

    if gps is not None:
        for const_name, key in [
            ("GPS_HISTORY", "history"), ("GPS_BENCHMARKS", "benchmarks"),
            ("GPS_BY_POSITION", "by_position"), ("GPS_REFERENCES", "references"),
        ]:
            html = re.sub(
                rf"const {const_name} = \[.*?\];",
                f"const {const_name} = " + json.dumps(gps[key], ensure_ascii=False) + ";",
                html, flags=re.S,
            )

    # valida antes de salvar
    for name, pat in [
        ("DATA", r"const DATA = (\[.*?\]);"),
        ("SD_TOTAL", r"const SD_TOTAL = (\{.*?\});"),
        ("REPORT_DATA", r"const REPORT_DATA = (\{.*?\});"),
        ("PLAYERS", r"const PLAYERS = (\[.*?\]);"),
        ("NPL_RANKINGS", r"const NPL_RANKINGS = (\{.*?\});"),
        ("GPS_HISTORY", r"const GPS_HISTORY = (\[.*?\]);"),
    ]:
        m = re.search(pat, html, re.S)
        if not m:
            raise RuntimeError(f"validação falhou: padrão '{name}' não encontrado no HTML")
        json.loads(m.group(1))

    HTML_FILE.write_text(html, encoding="utf-8")
    print(f"   HTML atualizado (backup em {HTML_FILE.with_suffix('.bak.html').name}).")
    print(f"   Rounds no painel: {sorted(set(r['Round'] for r in records))}")


def step5_update_gps():
    print("5) Atualizando dados de Treino (GPS)...")
    if not GPS_FILE.exists():
        print(f"   Aviso: {GPS_FILE.name} não encontrado, pulando esta etapa.")
        return None
    import openpyxl
    wb = openpyxl.load_workbook(GPS_FILE, data_only=True)
    ws = wb["GPS Data"]

    session_type = ws["B4"].value
    date_val = ws["E4"].value
    opponent = ws["H4"].value
    round_val = ws["K4"].value
    date_str = date_val.strftime("%Y-%m-%d") if hasattr(date_val, "strftime") else str(date_val)

    players = []
    r = 7
    while ws.cell(r, 1).value:
        position = ws.cell(r, 8).value or None
        if position and position.startswith("Attackers ("):
            position = "Attackers"
        players.append({
            "name": ws.cell(r, 1).value,
            "total_distance": ws.cell(r, 2).value,
            "hsr": ws.cell(r, 3).value,
            "max_speed": ws.cell(r, 4).value,
            "accelerations": ws.cell(r, 5).value,
            "decelerations": ws.cell(r, 6).value,
            "hmld": ws.cell(r, 7).value,
            "position": position,
            "detail": ws.cell(r, 9).value or None,
        })
        r += 1

    signature = f"{date_str}|{session_type}|{opponent or ''}"
    history = json.loads(GPS_HISTORY_FILE.read_text()) if GPS_HISTORY_FILE.exists() else []
    if any(s["signature"] == signature for s in history):
        print(f"   Essa sessão ('{signature}') já estava no histórico — não duplicada.")
    else:
        history.append({
            "signature": signature, "date": date_str, "type": session_type,
            "opponent": opponent, "round": round_val, "players": players,
        })
        history.sort(key=lambda s: s["date"])
        GPS_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"   Sessão de {date_str} ({session_type}) adicionada ao histórico "
              f"({len(history)} sessões no total).")

    def sheet_rows(name, max_rows=20):
        s = wb[name]
        rows = []
        for row in s.iter_rows(min_row=1, max_row=max_rows, values_only=True):
            if any(v is not None for v in row):
                rows.append([v for v in row if v is not None])
        return rows

    return {
        "history": history,
        "benchmarks": sheet_rows("GPS Benchmarks"),
        "by_position": sheet_rows("By Position"),
        "references": sheet_rows("References"),
    }


def step4_publish_to_github(round_num):
    print("4) Publicando no GitHub (site online)...")
    files = [
        "perth_azzurri_painel.html", "atualizar_painel.py", "extract_wyscout.py", "app.py",
        "npl_comparison_data.json", "team_stats_perth.xlsx", "Perth_SC.pdf",
        "NPL_Comparison.pdf", "perth_sc_last5_snapshot.json", "perth_sc_applied_matches.json",
        "Perth_SC_GPS_Weekly_Update_v3.xlsx", "gps_history.json",
    ]
    existing = [f for f in files if (BASE_DIR / f).exists()]
    subprocess.run(["git", "add"] + existing, cwd=BASE_DIR, check=True)
    result = subprocess.run(["git", "diff", "--cached", "--quiet"], cwd=BASE_DIR)
    if result.returncode == 0:
        print("   Nada novo para publicar (site já está atualizado).")
        return
    subprocess.run(
        ["git", "commit", "-m", f"Atualização semanal — Round {round_num}"],
        cwd=BASE_DIR, check=True,
    )
    push = subprocess.run(["git", "push"], cwd=BASE_DIR, capture_output=True, text=True)
    if push.returncode != 0:
        print("   AVISO: não consegui enviar para o GitHub automaticamente.")
        print("   " + push.stderr.strip().replace("\n", "\n   "))
        print("   Rode manualmente: git push")
    else:
        print("   Publicado! O site deve atualizar em 1-2 minutos.")


if __name__ == "__main__":
    dd = step1_update_excel()
    step2_update_npl_json(dd)
    img1, img2 = step2b_update_shots_images()
    players = step2c_update_players(dd)
    npl_flanks_img, npl_goals_img, npl_corners_img = step2d_update_npl_images()
    gps = step5_update_gps()
    step3_inject_html(dd, img1, img2, players, gps, npl_flanks_img, npl_goals_img, npl_corners_img)
    step4_publish_to_github(int(dd["Round"].max()))
    print("\nPronto! Abra perth_azzurri_painel.html no navegador para confirmar.")
